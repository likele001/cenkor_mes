"""订单 Excel 导入：页面填写订单头，Excel 仅含产品/型号/数量等明细。"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.crud.customer import get_customer_by_id
from app.crud.order import create_order, get_order_by_code
from app.crud.process_price import create_price, get_price_by_sku_process
from app.crud.process_route import (
    clone_default_route_to_product,
    get_default_route_for_product,
    get_first_default_route_template,
)
from app.crud.product import create_product, get_product_by_code, get_product_by_name
from app.crud.sku import create_sku, find_sku_by_attrs, get_sku_by_code
from app.models.customer import Customer
from app.models.product import Product
from app.models.sku import Sku
from app.services.code_generator import BizType, allocate_code

# Excel 明细表头别名
_DETAIL_HEADER_ALIASES: dict[str, set[str]] = {
    "seq": {"seq", "序号", "serial", "no"},
    "product_name": {"product_name", "产品名称", "产品", "品名"},
    "sku_name": {"sku_name", "型号名称", "型号", "规格型号"},
    "color": {"color", "颜色"},
    "material": {"material", "材料", "材质"},
    "spec": {"spec", "规格", "尺寸"},
    "qty": {"qty", "数量"},
    "line_remark": {"line_remark", "行备注", "备注", "明细备注"},
}


@dataclass
class OrderImportParams:
    customer_id: int
    order_name: str
    due_date: date | None = None
    remark: str | None = None
    order_code: str | None = None
    auto_create_product: bool = True
    auto_create_sku: bool = True
    default_unit_price: Decimal = Decimal("1")


@dataclass
class DetailLine:
    row: int
    product_name: str
    sku_name: str | None = None
    color: str | None = None
    material: str | None = None
    spec: str | None = None
    qty: int = 0
    line_remark: str | None = None


@dataclass
class ImportContext:
    products: dict[str, Product] = field(default_factory=dict)
    skus: dict[str, Sku] = field(default_factory=dict)
    warnings: list[dict[str, Any]] = field(default_factory=list)


def _norm_header(val: Any) -> str:
    return str(val).strip().lower().replace(" ", "_")


def _cell_str(row: tuple, col: int | None) -> str | None:
    if col is None or col >= len(row):
        return None
    v = row[col]
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _parse_qty(raw: Any) -> int | None:
    if raw is None:
        return None
    try:
        n = int(float(str(raw).strip()))
        return n if n >= 1 else None
    except (ValueError, TypeError):
        return None


def _split_product_sku_name(raw: str) -> tuple[str, str | None]:
    """支持「产品名称-型号名称」合并写法，按第一个 '-' 拆分。"""
    s = raw.strip()
    if "-" in s:
        product_part, sku_part = s.split("-", 1)
        product_part = product_part.strip()
        sku_part = sku_part.strip() or None
        if product_part:
            return product_part, sku_part
    return s, None


def build_col_map(header: tuple, aliases: dict[str, set[str]]) -> dict[str, int]:
    col_map: dict[str, int] = {}
    for col_idx, val in enumerate(header):
        if val is None:
            continue
        key = _norm_header(val)
        for logical, names in aliases.items():
            if key in names and logical not in col_map:
                col_map[logical] = col_idx
    return col_map


def parse_detail_lines(raw: bytes) -> list[DetailLine]:
    wb = load_workbook(BytesIO(raw), read_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Excel 文件为空")
    rows_iter = iter(ws.iter_rows(values_only=True))
    try:
        header = next(rows_iter)
    except StopIteration:
        raise ValueError("Excel 文件无数据行")
    col_map = build_col_map(header, _DETAIL_HEADER_ALIASES)
    if "qty" not in col_map:
        raise ValueError('缺少必要列「数量」')
    if "product_name" not in col_map:
        raise ValueError('缺少必要列「产品名称」')

    lines: list[DetailLine] = []
    for row_idx, row in enumerate(rows_iter, start=2):
        qty = _parse_qty(row[col_map["qty"]] if col_map["qty"] < len(row) else None)
        if qty is None:
            continue
        product_raw = _cell_str(row, col_map.get("product_name"))
        if not product_raw:
            continue
        sku_raw = _cell_str(row, col_map.get("sku_name"))
        product_name, sku_from_dash = _split_product_sku_name(product_raw)
        sku_name = sku_raw or sku_from_dash
        lines.append(
            DetailLine(
                row=row_idx,
                product_name=product_name,
                sku_name=sku_name,
                color=_cell_str(row, col_map.get("color")),
                material=_cell_str(row, col_map.get("material")),
                spec=_cell_str(row, col_map.get("spec")),
                qty=qty,
                line_remark=_cell_str(row, col_map.get("line_remark")),
            )
        )
    return lines


def build_import_template_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "订单明细"
    ws.append(["序号", "产品名称", "型号名称", "颜色", "材料", "规格", "数量", "行备注"])
    ws.append([1, "示例产品", "红色款", "红色", "塑料", "100x50", 100, ""])
    ws.append([2, "示例产品-蓝色款", "", "蓝色", "塑料", "100x50", 50, "合并写法示例"])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _warn(ctx: ImportContext, row: int, message: str) -> None:
    ctx.warnings.append({"row": row, "message": message})


def resolve_product(
    db: Session,
    line: DetailLine,
    ctx: ImportContext,
    *,
    allow_create: bool,
) -> Product:
    cache_key = line.product_name
    if cache_key in ctx.products:
        return ctx.products[cache_key]
    p = get_product_by_name(db, name=line.product_name)
    if p is None and allow_create:
        code = allocate_code(db, BizType.PRODUCT)
        p = create_product(
            db,
            code=code,
            name=line.product_name,
            category=None,
            unit="件",
            description=None,
            is_active=True,
        )
        template = get_first_default_route_template(db)
        if not template:
            raise ValueError("无可克隆的默认工艺路线，请先配置产品工艺路线")
        clone_default_route_to_product(db, p.id, template)
        _warn(ctx, line.row, f"已自动创建产品 {p.code}（{p.name}）并克隆默认工艺路线")
    if p is None:
        raise ValueError(f"产品「{line.product_name}」不存在，请勾选自动创建产品或先在系统中维护")
    ctx.products[cache_key] = p
    return p


def ensure_sku_process_prices(
    db: Session,
    sku_id: int,
    product_id: int,
    unit_price: Decimal,
) -> int:
    route = get_default_route_for_product(db, product_id)
    created = 0
    for step in route.steps:
        if get_price_by_sku_process(db, sku_id, step.process_id):
            continue
        create_price(db, sku_id=sku_id, process_id=step.process_id, unit_price=unit_price, is_active=True)
        created += 1
    return created


def resolve_sku(
    db: Session,
    product: Product,
    line: DetailLine,
    ctx: ImportContext,
    *,
    allow_create: bool,
    default_unit_price: Decimal,
) -> Sku:
    cache_key = f"{product.id}:{line.sku_name}:{line.color or ''}:{line.material or ''}:{line.spec or ''}"
    if cache_key in ctx.skus:
        return ctx.skus[cache_key]

    sku = find_sku_by_attrs(
        db,
        product_id=product.id,
        name=line.sku_name,
        color=line.color,
        material=line.material,
        spec=line.spec,
    )
    if sku is None and allow_create:
        code = allocate_code(db, BizType.SKU)
        while get_sku_by_code(db, code=code):
            code = allocate_code(db, BizType.SKU)
        sku = create_sku(
            db,
            product_id=product.id,
            code=code,
            name=line.sku_name,
            color=line.color,
            material=line.material,
            spec=line.spec,
            remark=None,
            is_active=True,
        )
        n = ensure_sku_process_prices(db, sku.id, product.id, default_unit_price)
        _warn(ctx, line.row, f"已自动创建型号 {sku.code}（{sku.name}），编号由系统生成，并创建 {n} 条工序工价")
    if sku is None:
        raise ValueError(f"型号「{line.sku_name}」在产品「{product.name}」下不存在，请勾选自动创建型号或先维护型号")
    if not sku.is_active:
        raise ValueError(f"型号 {sku.code} 已停用")
    ctx.skus[cache_key] = sku
    return sku


def import_single_order_from_excel(
    db: Session,
    params: OrderImportParams,
    raw: bytes,
) -> dict[str, Any]:
    customer = get_customer_by_id(db, customer_id=params.customer_id)
    if not customer:
        raise ValueError("客户不存在")
    if not customer.is_active:
        raise ValueError("客户已停用")

    order_name = (params.order_name or "").strip()
    if not order_name:
        raise ValueError("请填写订单名称")

    try:
        lines = parse_detail_lines(raw)
    except ValueError as e:
        return {
            "orders_created": 0,
            "lines_success": 0,
            "errors": [{"row": 0, "message": str(e)}],
            "warnings": [],
            "created_orders": [],
        }

    if not lines:
        return {
            "orders_created": 0,
            "lines_success": 0,
            "errors": [{"row": 0, "message": "无有效明细行（需填写产品名称、型号与数量）"}],
            "warnings": [],
            "created_orders": [],
        }

    ctx = ImportContext()
    errors: list[dict[str, Any]] = []
    order_items: list[tuple[int, int, int, str | None]] = []
    lines_success = 0

    for idx, line in enumerate(lines, start=1):
        if not line.sku_name:
            errors.append({"row": line.row, "message": "请填写「型号名称」，或在产品名称中使用「产品-型号」格式"})
            continue
        try:
            product = resolve_product(
                db, line, ctx, allow_create=params.auto_create_product
            )
            sku = resolve_sku(
                db,
                product,
                line,
                ctx,
                allow_create=params.auto_create_sku,
                default_unit_price=params.default_unit_price,
            )
            order_items.append((idx, sku.id, line.qty, line.line_remark))
            lines_success += 1
        except ValueError as e:
            errors.append({"row": line.row, "message": str(e)})

    if not order_items:
        return {
            "orders_created": 0,
            "lines_success": 0,
            "errors": errors,
            "warnings": ctx.warnings,
            "created_orders": [],
        }

    manual_code = (params.order_code or "").strip() or None
    if manual_code and get_order_by_code(db, code=manual_code):
        return {
            "orders_created": 0,
            "lines_success": lines_success,
            "errors": errors + [{"row": 0, "message": f"订单号 '{manual_code}' 已存在"}],
            "warnings": ctx.warnings,
            "created_orders": [],
        }

    order_code = manual_code or allocate_code(db, BizType.ORDER)
    while get_order_by_code(db, code=order_code):
        order_code = allocate_code(db, BizType.ORDER)

    remark_parts = [order_name]
    if params.remark:
        remark_parts.append(params.remark.strip())
    order = create_order(
        db,
        customer_id=customer.id,
        code=order_code,
        due_date=params.due_date,
        remark=" | ".join(remark_parts),
        items=order_items,
    )

    return {
        "orders_created": 1,
        "lines_success": lines_success,
        "errors": errors,
        "warnings": ctx.warnings,
        "created_orders": [{"id": order.id, "code": order.code}],
    }
